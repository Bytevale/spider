import asyncio
from playwright.async_api import async_playwright
import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import Font, PatternFill, Alignment
from fontTools.ttLib import TTFont
from PIL import Image, ImageFont, ImageDraw
import ddddocr


class MaoyanFontDecoder:
    def __init__(self, font_path):
        """
        初始化字体解码器
        :param font_path: 本地字体文件路径(.woff/.woff2/.ttf)
        """
        self.ocr = ddddocr.DdddOcr(show_ad=False)
        self.font_map = self.parse_font(font_path)

    def parse_font(self, font_path):
        """解析本地字体文件"""
        font = TTFont(font_path)
        cmap = font.getBestCmap()
        key_maps = {}

        # 创建字体映射表
        for code, name in cmap.items():
            char = chr(code)
            # 生成字符图片用于OCR识别
            img = Image.new('RGB', (100, 100), color='white')
            draw = ImageDraw.Draw(img)
            try:
                draw.text(
                    xy=(10, 10),
                    text=char,
                    font=ImageFont.truetype(font_path, size=80),
                    fill='black'
                )
                # 使用OCR识别字符对应的实际数字
                value = self.ocr.classification(img)
                key_maps[char] = value
            except Exception as e:
                print(f"字符 {char} 识别失败: {e}")
                continue

        print("字体映射表:", key_maps)
        return key_maps

    def decode_text(self, text):
        """解码使用自定义字体的文本"""
        if not text or not self.font_map:
            return text

        decoded = []
        for char in text:
            decoded.append(self.font_map.get(char, char))
        return ''.join(decoded)


async def scrape_maoyan_box_office(font_decoder):
    """
    使用XPath爬取猫眼电影实时票房数据（带字体反爬处理）
    :param font_decoder: 字体解码器实例
    :return: 包含票房数据的列表
    """
    data_list = []

    async with async_playwright() as p:
        # 启动浏览器
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            viewport={'width': 1280, 'height': 800}
        )
        page = await context.new_page()

        try:
            print("正在访问猫眼电影票房页面...")
            await page.goto('https://piaofang.maoyan.com/dashboard', timeout=60000)

            # 等待票房数据加载
            await page.wait_for_selector('xpath=//div[@class="movielist"]//tbody/tr', timeout=30000)
            print("页面加载完成，开始提取数据...")

            # 获取所有电影行
            movie_rows = await page.query_selector_all('xpath=//div[@class="movielist"]//tbody/tr')

            for row in movie_rows:
                try:
                    # 提取影片名称
                    name_element = await row.query_selector('xpath=.//p[@class="moviename-name"]')
                    name = await name_element.inner_text() if name_element else "未知影片"

                    # 提取综合票房（需要解码）
                    box_office_element = await row.query_selector('xpath=./td[2]')
                    box_office = await box_office_element.inner_text() if box_office_element else "0"
                    box_office = font_decoder.decode_text(box_office)

                    # 提取票房占比（需要解码）
                    ratio_element = await row.query_selector('xpath=./td[3]')
                    ratio = await ratio_element.inner_text() if ratio_element else "0%"
                    ratio = font_decoder.decode_text(ratio)

                    # 提取排片场次
                    screenings_element = await row.query_selector('xpath=./td[4]')
                    screenings = await screenings_element.inner_text() if screenings_element else "0"

                    # 添加到数据列表
                    data_list.append({
                        '影片名称': name.strip(),
                        '综合票房(万)': box_office.strip(),
                        '票房占比': ratio.strip(),
                        '排片场次': screenings.strip(),
                        '爬取时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })

                    print(f"已获取: {name} - {box_office} - {ratio} - {screenings}")

                except Exception as e:
                    print(f"提取单个电影数据出错: {e}")
                    continue

            print(f"共获取到{len(data_list)}条票房数据")
            await page.screenshot(path='debug_screenshot.png', full_page=True)
            print("已保存页面截图到 debug_screenshot.png")

        except Exception as e:
            print(f"爬取过程中出错: {e}")
        finally:
            await browser.close()

    return data_list


def save_to_excel(data, filename=None):
    """保存数据到Excel（保持不变）"""
    if not data:
        print("没有数据可保存")
        return

    if not filename:
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'猫眼电影实时票房_{current_time}.xlsx'

    df = pd.DataFrame(data)
    os.makedirs('output', exist_ok=True)
    filepath = os.path.join('output', filename)

    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='实时票房')

            workbook = writer.book
            worksheet = writer.sheets['实时票房']

            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

            column_widths = {'A': 30, 'B': 15, 'C': 12, 'D': 15, 'E': 20}
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        print(f"数据已保存到 {filepath}")
    except Exception as e:
        print(f"保存文件出错: {e}")


async def main():
    # 初始化字体解码器（替换为您的字体文件路径）
    font_decoder = MaoyanFontDecoder("cat4.woff")  # 修改为您的字体文件路径

    # 爬取数据
    data = await scrape_maoyan_box_office(font_decoder)

    # 保存数据
    if data:
        save_to_excel(data)
    else:
        print("未获取到任何数据")


if __name__ == '__main__':
    asyncio.run(main())