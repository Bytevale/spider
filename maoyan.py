


import asyncio
from playwright.async_api import async_playwright
import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment
from fontTools.ttLib import TTFont
from PIL import Image, ImageFont, ImageDraw
import ddddocr
import re
import requests
import base64
from functools import lru_cache
import tempfile
import io

class MaoyanFontDecoder:
    def __init__(self):
        self.ocr = ddddocr.DdddOcr(show_ad=False)
        self.font_map = None
        self.headers = {
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }
        # 猫眼当前使用的字体映射表（需要定期更新）
        self.hardcoded_map = {
            '\ue44d': '0', '\ue8ec': '1', '\ue6f5': '2',
            '\ue8f0': '3', '\ue6f6': '4', '\ue8ee': '5',
            '\ue8ed': '6', '\ue8eb': '7', '\ue8ef': '8',
            '\ue8ea': '9'
        }

    async def get_font_maps(self, page, max_retries=3):
        """通过分析网络请求获取动态加载的字体"""
        for attempt in range(max_retries):
            try:
                # 等待字体文件加载
                async with page.expect_response(lambda response: '.woff' in response.url) as response_info:
                    await page.evaluate('''() => {
                        document.fonts.ready.then(() => console.log('Fonts loaded'));
                    }''')

                response = await response_info.value
                font_url = response.url
                print(f"捕获到动态加载的字体URL: {font_url}")

                # 使用内存处理字体文件
                font_data = await response.body()
                font_file = io.BytesIO(font_data)
                self.font_map = self.parse_font_from_bytes(font_file)
                return

            except Exception as e:
                print(f"尝试 {attempt + 1}/{max_retries} 捕获字体失败: {str(e)}")
                if attempt == max_retries - 1:
                    print("使用备用字体映射")
                    self.font_map = self.hardcoded_map
                else:
                    await asyncio.sleep(1)

    def parse_font_from_bytes(self, font_file):
        """从内存中解析字体文件"""
        with tempfile.NamedTemporaryFile(suffix='.woff', delete=False) as temp_file:
            temp_file.write(font_file.getvalue())
            temp_file_path = temp_file.name

        try:
            font_map = self.parse_font(temp_file_path)
        finally:
            # 确保删除临时文件
            try:
                os.unlink(temp_file_path)
            except:
                pass

        return font_map

    @lru_cache(maxsize=1)
    def parse_font(self, font_path):
        """解析本地字体文件"""
        font = TTFont(font_path)
        cmap = font.getBestCmap()
        key_maps = {}

        for code, name in cmap.items():
            char = chr(code)
            img = Image.new('RGB', (100, 100), color='white')
            draw = ImageDraw.Draw(img)
            try:
                draw.text(
                    xy=(10, 10),
                    text=char,
                    font=ImageFont.truetype(font_path, size=80),
                    fill='black'
                )
                value = self.ocr.classification(img)
                key_maps[char] = value
            except Exception as e:
                print(f"字符 {char} 识别失败: {e}")
                continue

        print("字体映射表:", key_maps)
        return key_maps

    def decode_text(self, text):
        """解码使用自定义字体的文本"""
        if not text or not self.font_map:
            return text

        return ''.join([self.font_map.get(char, char) for char in text])




async def scrape_maoyan_box_office(font_decoder):
    data_list = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            viewport={'width': 1280, 'height': 800}
        )
        page = await context.new_page()

        try:
            print("正在访问猫眼电影票房页面...")
            await page.goto('https://piaofang.maoyan.com/dashboard', timeout=60000)

            # 获取字体映射
            await font_decoder.get_font_maps(page)

            await page.wait_for_selector('xpath=//div[@class="movielist"]//tbody/tr', timeout=30000)
            movie_rows = await page.query_selector_all('xpath=//div[@class="movielist"]//tbody/tr')

            for row in movie_rows:
                try:
                    name_element = await row.query_selector('xpath=.//p[@class="moviename-name"]')
                    name = await name_element.inner_text() if name_element else "未知影片"

                    box_office_element = await row.query_selector('xpath=./td[2]')
                    box_office = await box_office_element.inner_text() if box_office_element else "0"
                    box_office = font_decoder.decode_text(box_office)

                    ratio_element = await row.query_selector('xpath=./td[3]')
                    ratio = await ratio_element.inner_text() if ratio_element else "0%"
                    ratio = font_decoder.decode_text(ratio)

                    screenings_element = await row.query_selector('xpath=./td[4]')
                    screenings = await screenings_element.inner_text() if screenings_element else "0"

                    data_list.append({
                        '影片名称': name.strip(),
                        '综合票房(万)': box_office.strip(),
                        '票房占比': ratio.strip(),
                        '排片场次': screenings.strip(),
                        '爬取时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })

                except Exception as e:
                    print(f"提取单个电影数据出错: {e}")
                    continue

            await page.screenshot(path='debug_screenshot.png', full_page=True)
        finally:
            await context.close()
            await browser.close()

    return data_list


def save_to_excel(data, filename=None):
    """保存数据到Excel"""
    if not data:
        print("没有数据可保存")
        return

    if not filename:
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'猫眼电影实时票房_{current_time}.xlsx'

    df = pd.DataFrame(data)
    os.makedirs('output', exist_ok=True)
    filepath = os.path.join('output', filename)

    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='实时票房')
            workbook = writer.book
            worksheet = writer.sheets['实时票房']

            header_font = ExcelFont(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

            for col, width in {'A': 30, 'B': 15, 'C': 12, 'D': 15, 'E': 20}.items():
                worksheet.column_dimensions[col].width = width

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        print(f"数据已保存到 {filepath}")
    except Exception as e:
        print(f"保存文件出错: {e}")


async def main():
    font_decoder = MaoyanFontDecoder()
    data = await scrape_maoyan_box_office(font_decoder)
    if data:
        save_to_excel(data)
    else:
        print("未获取到任何数据")


if __name__ == '__main__':
    asyncio.run(main())
